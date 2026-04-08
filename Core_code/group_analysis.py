"""
group_analysis.py — 细胞群体基因表达比较与热力图
=================================================
群体 DE → marker 收集 → 热力图/dotplot → Excel 导出
所有图片输出到 fig/ 目录
"""

import numpy as np
import pandas as pd
import scanpy as sc
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import networkx as nx
from scipy.sparse import issparse
import os
from core import (RANDOM_SEED, fig_path, make_fig, full_pipeline, PROJECT_ROOT,
                  assign_mapper_labels, ensure_fig_dir)

sc.settings.verbosity = 1
plt.style.use('seaborn-v0_8-paper')


# ============================================================
# 群体分配
# ============================================================
def assign_cells_to_groups(graph, G, n_cells):
    comps = sorted(nx.connected_components(G), key=len, reverse=True)
    cell_comp = np.full(n_cells, -1, dtype=int)
    cell_node = np.array([''] * n_cells, dtype=object)
    cell_node_size = np.full(n_cells, np.inf)
    for ci, comp in enumerate(comps):
        for nid in comp:
            members = graph['nodes'][nid]
            size = len(members)
            for idx in members:
                if idx < n_cells:
                    cell_comp[idx] = ci
                    if size < cell_node_size[idx]:
                        cell_node_size[idx] = size
                        cell_node[idx] = nid
    labels = np.array([f'Group_{c}' if c >= 0 else 'Unassigned' for c in cell_comp])
    return labels, cell_comp, cell_node


# ============================================================
# DE & Markers
# ============================================================
def run_group_de(adata, group_key='cell_group'):
    sc.tl.rank_genes_groups(adata, group_key, method='wilcoxon',
                             use_raw=True, pts=True)
    return adata


def collect_top_markers(adata, group_key='cell_group', n_genes=15):
    groups = adata.obs[group_key].cat.categories.tolist()
    all_markers, seen = [], set()
    for g in groups:
        df = sc.get.rank_genes_groups_df(adata, group=g)
        df = df[df['pvals_adj'] < 0.05].head(n_genes * 2)
        for _, row in df.iterrows():
            gene = row['names']
            if gene not in seen:
                all_markers.append({'group': g, 'gene': gene,
                                     'logFC': row['logfoldchanges'],
                                     'padj': row['pvals_adj']})
                seen.add(gene)
            if sum(1 for m in all_markers if m['group'] == g) >= n_genes:
                break
    return pd.DataFrame(all_markers), list(seen)


def compute_group_mean(adata, group_key, genes):
    if adata.raw is not None:
        X = adata.raw[:, genes].X
        if issparse(X): X = X.toarray()
        df = pd.DataFrame(X, columns=genes, index=adata.obs_names)
    else:
        X = adata[:, genes].X
        if issparse(X): X = X.toarray()
        df = pd.DataFrame(X, columns=genes, index=adata.obs_names)
    df['group'] = adata.obs[group_key].values
    return df.groupby('group').mean()


# ============================================================
# 热力图
# ============================================================
def plot_heatmap(mean_df, marker_df, save_path=None):
    gene_order = [g for g in marker_df['gene'].tolist() if g in mean_df.columns]
    plot_df = mean_df[gene_order]
    zscore_df = (plot_df - plot_df.mean()) / (plot_df.std() + 1e-8)
    groups = sorted(mean_df.index.tolist())
    zscore_df = zscore_df.loc[groups]

    n_genes, n_groups = len(gene_order), len(groups)
    fig_width = max(16, n_genes * 0.35)
    fig_height = max(4, n_groups * 0.8 + 3)

    fig = plt.figure(figsize=(fig_width, fig_height))
    gs = gridspec.GridSpec(2, 2, height_ratios=[0.3, 10],
                           width_ratios=[10, 0.5], hspace=0.02, wspace=0.03)

    # 上方色条
    ax_top = fig.add_subplot(gs[0, 0])
    group_colors = plt.cm.Set3(np.linspace(0, 1, n_groups))
    gene_group_map = {}
    for _, row in marker_df.iterrows():
        if row['gene'] in gene_order:
            gene_group_map[row['gene']] = row['group']
    color_bar = np.zeros((1, n_genes, 4))
    for j, gene in enumerate(gene_order):
        g = gene_group_map.get(gene, groups[0])
        gi = groups.index(g) if g in groups else 0
        color_bar[0, j] = group_colors[gi]
    ax_top.imshow(color_bar, aspect='auto')
    ax_top.set_xticks([]); ax_top.set_yticks([])
    ax_top.set_ylabel('Marker\ngroup', fontsize=8, rotation=0, ha='right', va='center')

    # 主热力图
    ax_main = fig.add_subplot(gs[1, 0])
    vmax = min(zscore_df.values.max(), 4)
    vmin = max(zscore_df.values.min(), -4)
    im = ax_main.imshow(zscore_df.values, aspect='auto', cmap='RdBu_r',
                         vmin=vmin, vmax=vmax, interpolation='nearest')
    ax_main.set_yticks(range(n_groups))
    ax_main.set_yticklabels(groups, fontsize=10, fontweight='bold')
    ax_main.set_xticks(range(n_genes))
    ax_main.set_xticklabels(gene_order, rotation=90, fontsize=7, ha='center')

    # 分隔线
    prev_group = None
    for j, gene in enumerate(gene_order):
        g = gene_group_map.get(gene)
        if prev_group is not None and g != prev_group:
            ax_main.axvline(j - 0.5, color='black', linewidth=0.8, linestyle='--', alpha=0.5)
            ax_top.axvline(j - 0.5, color='black', linewidth=0.8)
        prev_group = g

    ax_cb = fig.add_subplot(gs[1, 1])
    plt.colorbar(im, cax=ax_cb, label='Z-score')
    fig.suptitle('Cell Group × Marker Gene Expression Heatmap (Z-scored)',
                 fontsize=14, fontweight='bold', y=0.98)

    if save_path:
        plt.savefig(save_path, dpi=200, bbox_inches='tight', facecolor='white')
    plt.close()


# ============================================================
# Excel 导出
# ============================================================
def export_excel(adata, group_key, mean_df, marker_df, cell_node, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    thin = Border(*(Side(style='thin', color='CCCCCC'),) * 4)
    hf = Font(name='Arial', bold=True, size=11)
    hfill = PatternFill('solid', fgColor='D6EAF8')
    df_ = Font(name='Arial', size=10)

    def style_hdr(ws, row=1):
        for c in ws[row]:
            c.font, c.fill, c.border = hf, hfill, thin
            c.alignment = Alignment(horizontal='center')

    # Sheet 1: 细胞归属
    ws1 = wb.active; ws1.title = 'Cell_Assignments'
    ws1.append(['Cell_ID', 'Group', 'Mapper_Node', 'Total_Counts', 'N_Genes'])
    style_hdr(ws1)
    obs = adata.obs
    for i, cid in enumerate(adata.obs_names):
        ws1.append([
            cid, str(obs[group_key].values[i]),
            str(cell_node[i]) if i < len(cell_node) else '',
            float(obs['total_counts'].values[i]) if 'total_counts' in obs else 0,
            int(obs['n_genes_by_counts'].values[i]) if 'n_genes_by_counts' in obs else 0,
        ])
        for c in ws1[ws1.max_row]: c.font, c.border = df_, thin
    for col, w in [('A', 18), ('B', 12), ('C', 20), ('D', 15), ('E', 12)]:
        ws1.column_dimensions[col].width = w
    ws1.auto_filter.ref = ws1.dimensions

    # Sheet 2: 平均表达
    ws2 = wb.create_sheet('Group_Mean_Expression')
    ws2.append(['Group'] + list(mean_df.columns))
    style_hdr(ws2)
    for gn in mean_df.index:
        ws2.append([gn] + [round(float(v), 4) for v in mean_df.loc[gn]])
        for c in ws2[ws2.max_row]: c.font, c.border = df_, thin
    ws2.column_dimensions['A'].width = 12
    ws2.freeze_panes = 'B2'

    # Sheet 3: Marker 基因
    ws3 = wb.create_sheet('Marker_Genes')
    ws3.append(['Group', 'Gene', 'Log2FC', 'Padj'])
    style_hdr(ws3)
    palette = ['FDEBD0', 'D5F5E3', 'D6EAF8', 'F5EEF8', 'FADBD8',
               'E8DAEF', 'D4EFDF', 'FCF3CF', 'F2D7D5', 'D5D8DC']
    gfills = {g: PatternFill('solid', fgColor=palette[i % len(palette)])
              for i, g in enumerate(marker_df['group'].unique())}
    for _, row in marker_df.iterrows():
        ws3.append([row['group'], row['gene'], round(float(row['logFC']), 3),
                     f"{float(row['padj']):.2e}"])
        fill = gfills.get(row['group'], PatternFill())
        for c in ws3[ws3.max_row]: c.font, c.border, c.fill = df_, thin, fill
    for col, w in [('A', 12), ('B', 18), ('C', 10), ('D', 14)]:
        ws3.column_dimensions[col].width = w
    ws3.auto_filter.ref = ws3.dimensions

    # Sheet 4: 统计摘要
    ws4 = wb.create_sheet('Group_Summary')
    ws4.append(['Group', 'N_Cells', 'Pct', 'Median_UMI', 'Median_Genes',
                'Top3_Markers', 'Suggested_Type'])
    style_hdr(ws4)
    total = adata.n_obs
    for g in sorted(adata.obs[group_key].cat.categories):
        mask = adata.obs[group_key] == g
        n = mask.sum()
        mu = np.median(obs['total_counts'].values[mask]) if 'total_counts' in obs else 0
        mg = np.median(obs['n_genes_by_counts'].values[mask]) if 'n_genes_by_counts' in obs else 0
        t3 = ', '.join(marker_df[marker_df['group'] == g]['gene'].head(3).tolist()) or 'N/A'
        ws4.append([g, int(n), round(n / total * 100, 1), round(float(mu)),
                     int(mg), t3, ''])
        for c in ws4[ws4.max_row]: c.font, c.border = df_, thin
    yellow = PatternFill('solid', fgColor='FFFFCC')
    for r in range(2, ws4.max_row + 1):
        ws4.cell(row=r, column=7).fill = yellow
    for col, w in [('A', 12), ('B', 10), ('C', 8), ('D', 12),
                    ('E', 12), ('F', 35), ('G', 25)]:
        ws4.column_dimensions[col].width = w

    wb.save(output_path)
    print(f"[Info] Excel: {output_path}")


# ============================================================
# Main
# ============================================================
if __name__ == "__main__":
    ensure_fig_dir('group')
    figp = make_fig('group')

    # 1. 全流程
    adata, graph, G, comps, mapper_labels, lens_1d = full_pipeline()

    # 2. 分配群体
    group_labels, cell_comp, cell_node = assign_cells_to_groups(graph, G, adata.n_obs)
    adata.obs['cell_group'] = pd.Categorical(group_labels)

    print("\nGroup sizes:")
    for g in sorted(adata.obs['cell_group'].cat.categories):
        n = (adata.obs['cell_group'] == g).sum()
        print(f"  {g}: {n} cells ({n / adata.n_obs * 100:.1f}%)")

    # 3. DE
    run_group_de(adata, 'cell_group')
    marker_df, marker_genes = collect_top_markers(adata, 'cell_group', n_genes=15)
    print(f"\n{len(marker_genes)} unique markers collected.")
    for g in sorted(marker_df['group'].unique()):
        gs = marker_df[marker_df['group'] == g]['gene'].head(5).tolist()
        print(f"  {g}: {', '.join(gs)}")

    # 4. 平均表达
    mean_df = compute_group_mean(adata, 'cell_group', marker_genes)
    print(f"\nExpression matrix: {mean_df.shape}")

    # 5. 热力图
    plot_heatmap(mean_df, marker_df, save_path=figp('heatmap.png'))
    print(f"[Info] Saved: {figp('heatmap.png')}")

    # 6. Dotplot
    try:
        marker_dict = {}
        for g in sorted(marker_df['group'].unique()):
            gs = marker_df[marker_df['group'] == g]['gene'].head(8).tolist()
            if adata.raw is not None:
                gs = [x for x in gs if x in adata.raw.var_names]
            marker_dict[g] = gs
        if marker_dict:
            sc.pl.dotplot(adata, var_names=marker_dict, groupby='cell_group',
                           use_raw=True, standard_scale='var', show=False)
            plt.savefig(figp('dotplot.png'), dpi=200, bbox_inches='tight')
            print(f"[Info] Saved: {figp('dotplot.png')}")
            plt.close()
    except Exception as e:
        print(f"[Warning] Dotplot: {e}")

    # 7. Matrixplot
    try:
        flat = []
        for g in sorted(marker_df['group'].unique()):
            gs = marker_df[marker_df['group'] == g]['gene'].head(5).tolist()
            if adata.raw is not None:
                gs = [x for x in gs if x in adata.raw.var_names]
            flat.extend(gs)
        if flat:
            sc.pl.matrixplot(adata, var_names=flat, groupby='cell_group',
                              use_raw=True, standard_scale='var', cmap='Blues', show=False)
            plt.savefig(figp('matrixplot.png'), dpi=200, bbox_inches='tight')
            print(f"[Info] Saved: {figp('matrixplot.png')}")
            plt.close()
    except Exception as e:
        print(f"[Warning] Matrixplot: {e}")

    # 8. Excel
    export_excel(adata, 'cell_group', mean_df, marker_df, cell_node,
                 output_path=os.path.join(PROJECT_ROOT, 'cell_groups_analysis.xlsx'))

    print(f"\n[Done] Figures in fig/, Excel in cell_groups_analysis.xlsx")
